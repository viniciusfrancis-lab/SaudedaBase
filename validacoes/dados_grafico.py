"""
dados_grafico.py
─────────────────────────────────────────────────────────────
Módulo para gerar métricas da base carregada via pygei.
Executa silenciosamente no app.py sem exibir nada na UI.
"""

import traceback
import pandas as pd
from pathlib import Path
from datetime import datetime
from openpyxl import load_workbook, Workbook


# Caminho absoluto - independente de onde o script é executado
BASE_DIR = Path(__file__).resolve().parent
PASTA_SAIDA = BASE_DIR / "dados_graficos"
ARQUIVO_SAIDA = PASTA_SAIDA / "dados_graficos.xlsx"


def checar_alunos_total(df: pd.DataFrame) -> int:
    """Retorna total de alunos únicos (prioriza id_aluno, fallback para nm_aluno)."""

    if df is None or df.empty:
        return 0

    if "id_aluno" in df.columns:
        return int(df["id_aluno"].nunique())

    if "nm_aluno" in df.columns:
        return int(df["nm_aluno"].nunique())

    return 0


def gerar_relatorio(df: pd.DataFrame) -> pd.DataFrame:
    """
    Atualiza dados_graficos.xlsx preservando colunas C e D (escritas pelo processar.py).
    Escreve apenas:
      - A1: "Funcao"        | B1: "Resultado"
      - A2: "checar_alunos_total" | B2: <total>
      - A10: data/hora da última atualização
    """

    total_alunos = checar_alunos_total(df)

    relatorio = pd.DataFrame({
        "Funcao": ["checar_alunos_total"],
        "Resultado": [total_alunos]
    })

    try:
        PASTA_SAIDA.mkdir(parents=True, exist_ok=True)

        # Abre o arquivo existente OU cria um novo - sem sobrescrever C/D
        if ARQUIVO_SAIDA.exists():
            wb = load_workbook(ARQUIVO_SAIDA)
            ws = wb.active
        else:
            wb = Workbook()
            ws = wb.active

        # Escreve apenas nas células A1, B1, A2, B2 e A10
        ws["A1"] = "Funcao"
        ws["B1"] = "Resultado"
        ws["A2"] = "checar_alunos_total"
        ws["B2"] = total_alunos

        data_hora = datetime.now().strftime("%d/%m/%Y %H:%M:%S")
        ws["A10"] = f"Última atualização: {data_hora}"

        wb.save(ARQUIVO_SAIDA)

    except PermissionError as erro:
        print(
            f"[dados_grafico] Arquivo bloqueado (Excel aberto ou OneDrive sincronizando?): {erro}"
        )
        traceback.print_exc()

    except Exception as erro:
        print(f"[dados_grafico] Erro ao salvar relatório: {erro}")
        traceback.print_exc()

    return relatorio


def executar_silenciosamente(df: pd.DataFrame) -> dict:
    """
    Executa tudo silenciosamente e retorna métricas em dict.
    """

    try:
        relatorio = gerar_relatorio(df)

        return {
            "alunos_total": int(
                relatorio.loc[
                    relatorio["Funcao"] == "checar_alunos_total",
                    "Resultado"
                ].iloc[0]
            )
        }

    except Exception as erro:
        print(f"[dados_grafico] Erro na execução silenciosa: {erro}")
        traceback.print_exc()

        return {
            "alunos_total": 0
        }
    
    #######################        
        
if __name__ == "__main__":

    caminho_parquet = (
        r"N:\04 - Relatórios\Relatórios SGE\13 - Relatórios JIRA"
        r"\03- Relatórios jiras base de alunos semanal"
        r"\2026\05 - Maio\ALUNO"
        r"\parquet_BASE_DE_ALUNO_28_05_2026.parquet"
    )

    print("Lendo parquet...")

    dfSeges = pd.read_parquet(caminho_parquet)

    print(f"Linhas originais: {len(dfSeges):,}")

    # =========================================================
    # FILTROS
    # =========================================================

    num_ano_letivo = [
        col
        for col in dfSeges.num_ano_letivo.unique()
        if '2026' in str(col)
    ]

    dfSeges = dfSeges[
        dfSeges.num_ano_letivo.isin(num_ano_letivo)
    ]

    dfSeges = dfSeges[
        dfSeges['num_ano_letivo'] != '2026 - MEPES'
    ]

    # OPCIONAIS
    # dfSeges = dfSeges[
    #     dfSeges['situacao_enturmacao'] == "Em curso"
    # ]

    # dfSeges = dfSeges[
    #     dfSeges['situacao_matricula'] == "Em curso"
    # ]

    print(f"Linhas após filtros: {len(dfSeges):,}")

    # =========================================================
    # EXECUÇÃO
    # =========================================================

    resultado = executar_silenciosamente(dfSeges)

    print("\nResultado:")
    print(resultado)