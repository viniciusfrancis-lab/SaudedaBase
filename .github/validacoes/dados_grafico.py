"""
dados_grafico.py
─────────────────────────────────────────────────────────────
Módulo para gerar métricas da base carregada via pygei.
Executa silenciosamente no app.py sem exibir nada na UI.
"""

import pandas as pd
from pathlib import Path
from datetime import datetime
from openpyxl import load_workbook


def checar_alunos_total(df: pd.DataFrame) -> int:
    """Retorna total de alunos únicos."""
    
    if "nm_aluno" not in df.columns:
        return 0

    return df["nm_aluno"].nunique()


def gerar_relatorio(df: pd.DataFrame) -> pd.DataFrame:

    resultados = {
        "Funcao": [
            "checar_alunos_total"
        ],

        "Resultado": [
            len(df["nm_aluno"].unique())
        ]
    }

    relatorio = pd.DataFrame(resultados)

    try:
        pasta_saida = Path("dados_graficos")
        pasta_saida.mkdir(parents=True, exist_ok=True)

        arquivo_saida = pasta_saida / "dados_graficos.xlsx"

        # Salva DataFrame
        relatorio.to_excel(
            arquivo_saida,
            index=False,
            engine="openpyxl"
        )

        # Abre o Excel novamente
        wb = load_workbook(arquivo_saida)
        ws = wb.active

        # Escreve data/hora na célula A10
        data_hora = datetime.now().strftime("%d/%m/%Y %H:%M:%S")

        ws["A10"] = f"Última atualização: {data_hora}"

        # Salva alterações
        wb.save(arquivo_saida)

    except Exception as erro:
        print(f"Erro ao salvar relatório: {erro}")

    return relatorio


def executar_silenciosamente(df: pd.DataFrame) -> dict:
    """
    Executa tudo silenciosamente e retorna métricas em dict.
    """

    try:
        relatorio = gerar_relatorio(df)

        return {
            "alunos_total": int(
                relatorio.loc[
                    relatorio["Funcao"] == "checar_alunos_total",
                    "Resultado"
                ].iloc[0]
            )
        }

    except Exception as erro:
        print(f"Erro na execução silenciosa: {erro}")

        return {
            "alunos_total": 0
        }