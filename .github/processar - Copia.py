"""
processar.py
------------
Roda o pipeline completo de validações e salva os resultados em parquet.
Execute manualmente ou agende via Task Scheduler do Windows.

Saídas:
  - resultados.parquet   → todas as linhas de alerta (para o app exibir)
  - resumo.parquet       → contagem por tipo de alerta (para o app exibir)
  - dados_graficos.xlsx  → colunas C e D atualizadas (para o Power BI / gráficos)
"""

import pandas as pd
from datetime import datetime
from pathlib import Path
from openpyxl import load_workbook

from carregador import aplicar_filtros_cpf, COLUNAS_OBRIGATORIAS
from validacoes.cpf import checar_cpf
from validacoes.matricula import checar_duplicatas
from validacoes.matricula_data import checar_alteracao_dt_matricula
from validacoes.flag_deficiencia import checar_flag_deficiencia
from validacoes.cor_raca import checar_sem_autodeclaracao_racial
from validacoes.dados_grafico import executar_silenciosamente
from pygei.seges import alunos

# ─── Caminhos ──────────────────────────────────────────────
BASE_DIR = Path(__file__).parent / "dados_graficos"

PARQUET_RESULTADOS = BASE_DIR / "resultados.parquet"
PARQUET_RESUMO     = BASE_DIR / "resumo.parquet"
CAMINHO_EXCEL      = BASE_DIR / "dados_graficos.xlsx"

# ─── Rótulos fixos de todos os alertas possíveis ───────────
TODOS_ALERTAS = [
    "CPF inválido/em branco",    
    "Matrícula duplicada",
    "Deficiência sem descrição",
    "Descrição de deficiência indevida",
    "dt_matricula alterada",
    "Matrícula retroativa",    
    "Mudança de id_aluno",
    "Frequência io-iô",
    "Sem_autodeclaracao_racial",    
]


def _tratar_base(df: pd.DataFrame) -> pd.DataFrame:
    df = df.copy()
    df.columns = df.columns.str.strip().str.lower()
    for col in df.columns:
        df[col] = df[col].astype(str).where(df[col].notna(), None)
    if "cpf" in df.columns:
        df["cpf"] = df["cpf"].str.strip()
    return aplicar_filtros_cpf(df)


def _salvar_excel(resumo: pd.DataFrame) -> None:
    """
    Atualiza colunas C e D do dados_graficos.xlsx com o resumo de alertas.
    Usa openpyxl para preservar o conteúdo existente (colunas A e B).
    """
    wb = load_workbook(CAMINHO_EXCEL)
    ws = wb.active

    # Cabeçalhos C1 e D1
    ws["C1"] = "col_2"
    ws["D1"] = "Erros"

    # Limpa C2:D20 antes de escrever
    for row in range(2, 21):
        ws.cell(row=row, column=3, value=None)
        ws.cell(row=row, column=4, value=None)

    # Preenche C2:D(n+1)
    for i, (alerta, qtd) in enumerate(zip(resumo["alerta"], resumo["Qtd"]), start=2):
        ws.cell(row=i, column=3, value=alerta)
        ws.cell(row=i, column=4, value=int(qtd))

    wb.save(CAMINHO_EXCEL)


def main():
    inicio = datetime.now()
    print(f"[{inicio:%H:%M:%S}] Iniciando processamento...")

    # 1. Carrega base via pygei
    print("  Carregando base SEGES via pygei...")
    df_raw = alunos.latest()
    print(f"  {len(df_raw):,} registros brutos carregados.")

    # 2. Trata base
    df = _tratar_base(df_raw)
    print(f"  {len(df):,} registros após filtros (ano letivo 2026, Em curso, Regular).")

    faltando = [c for c in COLUNAS_OBRIGATORIAS if c not in df.columns]
    if faltando:
        raise ValueError(f"Colunas ausentes na base: {faltando}")

    # 3. Execução silenciosa do dados_grafico.py
    executar_silenciosamente(df)

    # 4. Auditoria temporal
    print("  Executando auditoria temporal...")
    resultado_dt = checar_alteracao_dt_matricula()

    # 5. Todas as validações
    print("  Executando validações...")
    resultados = pd.concat([
        checar_cpf(df),        
        checar_duplicatas(df),
        checar_flag_deficiencia(df),
        checar_sem_autodeclaracao_racial(df),
        resultado_dt,
    ], ignore_index=True)

    # 6. Monta resumo
    contagem = (
        resultados.groupby("alerta").size().reset_index(name="Qtd")
        if not resultados.empty
        else pd.DataFrame(columns=["alerta", "Qtd"])
    )
    resumo = (
        pd.DataFrame({"alerta": TODOS_ALERTAS})
        .merge(contagem, on="alerta", how="left")
        .fillna({"Qtd": 0})
        .astype({"Qtd": int})
    )

    # 7. Salva parquets
    print("  Salvando parquets...")
    resultados.to_parquet(PARQUET_RESULTADOS, index=False)
    resumo.to_parquet(PARQUET_RESUMO, index=False)
    print(f"  ✅ resultados.parquet → {len(resultados):,} alertas")
    print(f"  ✅ resumo.parquet     → {len(resumo)} tipos de alerta")

    # 8. Atualiza Excel
    print("  Atualizando dados_graficos.xlsx...")
    _salvar_excel(resumo)
    print("  ✅ dados_graficos.xlsx atualizado")

    fim = datetime.now()
    print(f"[{fim:%H:%M:%S}] Processamento concluído em {(fim - inicio).seconds}s.")


if __name__ == "__main__":
    main()
