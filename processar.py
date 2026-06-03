"""
processar.py
------------
Roda o pipeline completo de validações e salva os resultados em parquet.
Execute manualmente ou agende via Task Scheduler do Windows.

RETOMÁVEL (idempotente):
  - Cada etapa salva seu próprio parquet em dados_graficos/temp/.
  - Ao rodar de novo, etapas cujo parquet já existe são PULADAS.
  - Se algum alerta deu erro, basta rodar de novo: só o que faltou é processado.
  - O temp só é apagado depois da consolidação final bem-sucedida.
  - Use  python processar.py --force  para ignorar o temp e refazer tudo do zero.

Saídas:
  - resultados.parquet   → todas as linhas de alerta (para o app exibir)
  - resumo.parquet       → contagem por tipo de alerta (para o app exibir)
  - dados_graficos.xlsx  → colunas C e D atualizadas (para o Power BI / gráficos)
"""

import sys
import pandas as pd
from datetime import datetime
from pathlib import Path
from openpyxl import load_workbook

from carregador import aplicar_filtros_cpf, COLUNAS_OBRIGATORIAS
#from alertas_config import TODOS_ALERTAS
from alertas_config import (
    TODOS_ALERTAS,
    ALERTA_DT_MATRICULA, ALERTA_RETROATIVA, ALERTA_MUDANCA_ID, ALERTA_IOIO,
)
from validacoes.alerta_ultima_aparicao import checar_ultima_aparicao, gerar_alerta
from validacoes.cpf import checar_cpf
from validacoes.matricula import checar_duplicatas
from validacoes.matricula_data import checar_alteracao_dt_matricula
from validacoes.flag_deficiencia import checar_flag_deficiencia
from validacoes.cor_raca import checar_sem_autodeclaracao_racial
from validacoes.dados_grafico import executar_silenciosamente
from pygei.seges import alunos

# ─── Flags de execução ─────────────────────────────────────
FORCAR = "--force" in sys.argv  # reprocessa tudo, ignorando o temp

# ─── Caminhos ──────────────────────────────────────────────
BASE_DIR = Path(__file__).parent / "dados_graficos"

PARQUET_RESULTADOS = BASE_DIR / "resultados.parquet"
PARQUET_RESUMO     = BASE_DIR / "resumo.parquet"
CAMINHO_EXCEL      = BASE_DIR / "dados_graficos.xlsx"
SAIDA_ULTIMA_APARICAO = BASE_DIR / "alerta_ultima_aparicao.xlsx"

TEMP_DIR = BASE_DIR / "temp"
ARQ_AUDITORIA = TEMP_DIR / "todas_regras_auditoria.parquet"

# Colunas que costumam vir com tipo misto (int/str, date/NaN) entre as
# diferentes validações. Convertidas para string antes de salvar em parquet,
# evitando ArrowTypeError no momento do to_parquet.
COLUNAS_FORCAR_STRING = [
    # IDs e códigos
    "id_aluno", "cpf", "inep_escola", "ra_aluno", "id_turma",
    # Datas (algumas validações entregam datetime.date, outras string ou NaN)
    "ultima_aparicao", "primeira_aparicao",
    "dt_matricula", "data_nascimento", "data_referencia",
]


def _tratar_base(df: pd.DataFrame) -> pd.DataFrame:
    df = df.copy()
    df.columns = df.columns.str.strip().str.lower()
    for col in df.columns:
        df[col] = df[col].astype(str).where(df[col].notna(), None)
    if "cpf" in df.columns:
        df["cpf"] = df["cpf"].str.strip()
    return aplicar_filtros_cpf(df)


def _normalizar_tipos(df: pd.DataFrame) -> pd.DataFrame:
    """
    Garante que colunas que podem vir com tipo misto entre validações
    diferentes (int vs str, datetime.date vs NaN) saiam como string.
    Evita ArrowTypeError no momento de salvar o parquet.
    """
    for col in COLUNAS_FORCAR_STRING:
        if col in df.columns:
            df[col] = df[col].astype("string")
    return df


def _ja_existe(caminho: Path) -> bool:
    """True se o parquet já existe no temp (e --force não foi usado)."""
    return not FORCAR and caminho.exists()


def _etapa(nome: str, arquivo: str, func, *args, **kwargs) -> None:
    """
    Roda uma validação SOMENTE se o parquet correspondente ainda não existe
    no temp. Caso já exista, pula. Salva o resultado normalizado em parquet.
    """
    destino = TEMP_DIR / arquivo
    if _ja_existe(destino):
        print(f"  ✓ {nome:<24} (já existe, pulando)")
        return
    print(f"  → {nome}...")
    resultado = func(*args, **kwargs)
    print(f"     {len(resultado):,} alertas")
    _normalizar_tipos(resultado).to_parquet(destino, index=False)
    del resultado  # ✅ LIBERA MEMÓRIA


def _precisa_da_base() -> bool:
    """
    A base (df) só precisa ser carregada/tratada se ALGUMA etapa que depende
    dela ainda não foi salva. Se todas já existem, pulamos o carregamento caro.
    """
    if FORCAR:
        return True
    etapas_que_usam_df = [
        "cpf.parquet", "duplicatas.parquet",
        "deficiencia.parquet", "cor_raca.parquet",
    ]
    return any(not (TEMP_DIR / a).exists() for a in etapas_que_usam_df)


def _salvar_excel(resumo: pd.DataFrame) -> None:
    """
    Atualiza colunas C e D do dados_graficos.xlsx com o resumo de alertas.
    Se o arquivo já existe, preserva o conteúdo existente (colunas A e B).
    Se NÃO existe, cria um novo workbook do zero.
    """
    if CAMINHO_EXCEL.exists():
        wb = load_workbook(CAMINHO_EXCEL)
        ws = wb.active
    else:
        from openpyxl import Workbook
        print(f"     ℹ️  {CAMINHO_EXCEL.name} não existe — criando do zero.")
        CAMINHO_EXCEL.parent.mkdir(parents=True, exist_ok=True)
        wb = Workbook()
        ws = wb.active

    # Cabeçalhos C1 e D1
    ws["C1"] = "col_2"
    ws["D1"] = "Erros"

    # Limpa C2:D20 antes de escrever
    for row in range(2, 21):
        ws.cell(row=row, column=3, value=None)
        ws.cell(row=row, column=4, value=None)

    # Preenche C2:D(n+1)
    for i, (alerta, qtd) in enumerate(zip(resumo["alerta"], resumo["Qtd"]), start=2):
        ws.cell(row=i, column=3, value=alerta)
        ws.cell(row=i, column=4, value=int(qtd))

    wb.save(CAMINHO_EXCEL)


def main():
    inicio = datetime.now()
    print(f"[{inicio:%H:%M:%S}] Iniciando processamento...")
    if FORCAR:
        print("  ⚠️  --force ativo: ignorando temp e reprocessando tudo.")

    # Garante que o temp existe ANTES de qualquer etapa
    TEMP_DIR.mkdir(parents=True, exist_ok=True)

    # ────────────────────────────────────────────────────────────────────────
    # 1-3. CARREGA / TRATA BASE (só se alguma etapa que usa df estiver faltando)
    # ────────────────────────────────────────────────────────────────────────
    df = None
    if _precisa_da_base():
        print("  Carregando base SEGES via pygei...")
        df_raw = alunos.latest()
        print(f"  {len(df_raw):,} registros brutos carregados.")

        df = _tratar_base(df_raw)
        print(f"  {len(df):,} registros após filtros (ano letivo 2026, Em curso, Regular).")
        del df_raw  # ✅ LIBERA MEMÓRIA

        faltando = [c for c in COLUNAS_OBRIGATORIAS if c not in df.columns]
        if faltando:
            raise ValueError(f"Colunas ausentes na base: {faltando}")

        # Execução silenciosa do dados_grafico.py (só faz sentido com a base em mãos)
        print("  Executando dados_grafico.py (silencioso)...")
        executar_silenciosamente(df)
    else:
        print("  ✓ Todas as etapas que usam a base já existem — pulando carregamento.")

    # ────────────────────────────────────────────────────────────────────────
    # 4. VALIDAÇÕES GERAIS (retomáveis)
    # ────────────────────────────────────────────────────────────────────────
    print("\n  ─── VALIDAÇÕES GERAIS ───")
    _etapa("CPF",                  "cpf.parquet",         checar_cpf, df)
    _etapa("Duplicatas",           "duplicatas.parquet",  checar_duplicatas, df)
    _etapa("Flag Deficiência",     "deficiencia.parquet", checar_flag_deficiencia, df)
    _etapa("Autodeclaração Racial","cor_raca.parquet",    checar_sem_autodeclaracao_racial, df)

    del df  # ✅ LIBERA MEMÓRIA (pode ser None, sem problema)

    # ────────────────────────────────────────────────────────────────────────
    # 5. REGRAS DE AUDITORIA (retomável)
    # ────────────────────────────────────────────────────────────────────────
    print("\n  ─── REGRAS DE AUDITORIA ───")

    if _ja_existe(ARQ_AUDITORIA):
        print("  ✓ todas_regras_auditoria.parquet (já existe, pulando processamento)")
        df_regras_auditoria = pd.read_parquet(ARQ_AUDITORIA)
        print(f"     {len(df_regras_auditoria):,} alertas lidos do disco")
    else:
        print("  → Executando todas as regras de auditoria (matricula_data)...")
        resultado_todas_regras = checar_alteracao_dt_matricula()

        print(f"  ✅ checar_alteracao_dt_matricula() concluído!")
        print(f"     Total de linhas: {len(resultado_todas_regras):,}")

        if not resultado_todas_regras.empty:
            print(f"     Colunas disponíveis: {list(resultado_todas_regras.columns)}")

            # Salva TODAS as regras juntas primeiro (garantia / fonte da verdade)
            _normalizar_tipos(resultado_todas_regras).to_parquet(ARQ_AUDITORIA, index=False)
            print(f"  ✅ Salvo: todas_regras_auditoria.parquet")

            # Separa por tipo de alerta (se a coluna existir)
            if "alerta" in resultado_todas_regras.columns:
                alertas_unicos = resultado_todas_regras["alerta"].unique()
                print(f"     Tipos de alerta encontrados: {len(alertas_unicos)}")

                mapeamento_regras = {
                    ALERTA_RETROATIVA: "regra1_matricula_retroativa.parquet",
                    ALERTA_IOIO: "regra2_frequencia_ioio.parquet",
                    ALERTA_MUDANCA_ID: "regra3a_mudanca_id.parquet",
                    ALERTA_DT_MATRICULA: "regra4_alteracao_dt_matricula.parquet",
                }
                #mapeamento_regras = {
                 #   "Matrícula retroativa": "regra1_matricula_retroativa.parquet",
                  #  "Frequência io-iô": "regra2_frequencia_ioio.parquet",
                   # "Mudança de id_aluno": "regra3a_mudanca_id.parquet",
                    #"Unificação de ID (mesmo nasc)": "regra3b_unificacao_id.parquet",
                    #"Alteração em dt_matricula": "regra4_alteracao_dt_matricula.parquet",
                    #"Cronologia inválida": "regra5_cronologia_invalida.parquet",
                #}

                for alerta_nome, arquivo in mapeamento_regras.items():
                    df_regra = resultado_todas_regras[resultado_todas_regras["alerta"] == alerta_nome]
                    if not df_regra.empty:
                        _normalizar_tipos(df_regra).to_parquet(TEMP_DIR / arquivo, index=False)
                        print(f"     • {alerta_nome:<40} → {len(df_regra):>6,} alertas")
            else:
                print("  ⚠️  Coluna 'alerta' não encontrada, salvando tudo junto apenas.")
        else:
            print("  ℹ️  Nenhuma regra de auditoria gerada (DataFrame vazio).")

        df_regras_auditoria = resultado_todas_regras.copy()
        del resultado_todas_regras  # ✅ LIBERA MEMÓRIA

    print(f"  → Total regras de auditoria: {len(df_regras_auditoria):,} alertas")
    if not df_regras_auditoria.empty and "id_aluno" in df_regras_auditoria.columns:
        print(f"     ({df_regras_auditoria['id_aluno'].nunique():,} alunos únicos)")

    # ────────────────────────────────────────────────────────────────────────
    # 6. ÚLTIMA APARIÇÃO (retomável — depende das regras de auditoria)
    # ────────────────────────────────────────────────────────────────────────
    print("\n  ─── ÚLTIMA APARIÇÃO (FILTRADO) ───")
    arq_ultima = TEMP_DIR / "ultima_aparicao.parquet"

    if _ja_existe(arq_ultima):
        print("  ✓ ultima_aparicao.parquet (já existe, pulando)")
    else:
        print("  → Gerando alerta de última aparição...")
        # ✅ Passa APENAS as regras de auditoria para exclusão
        resultado_ultima_aparicao = checar_ultima_aparicao(
            df_regras_auditoria=df_regras_auditoria if not df_regras_auditoria.empty else None
        )
        print(f"     {len(resultado_ultima_aparicao):,} alertas")
        if not resultado_ultima_aparicao.empty and "id_aluno" in resultado_ultima_aparicao.columns:
            print(f"     ({resultado_ultima_aparicao['id_aluno'].nunique():,} alunos únicos)")
            print("     (excluídos apenas alunos das regras de auditoria)")

        _normalizar_tipos(resultado_ultima_aparicao).to_parquet(arq_ultima, index=False)
        del resultado_ultima_aparicao  # ✅ LIBERA MEMÓRIA

        # Gera também o Excel standalone de última aparição
        print("  → Gerando alerta_ultima_aparicao.xlsx...")
        df_alerta_xlsx = gerar_alerta(verbose=False)
        if not df_alerta_xlsx.empty:
            SAIDA_ULTIMA_APARICAO.parent.mkdir(parents=True, exist_ok=True)
            df_alerta_xlsx.to_excel(SAIDA_ULTIMA_APARICAO, index=False)
            print(f"     ✅ alerta_ultima_aparicao.xlsx atualizado ({len(df_alerta_xlsx):,} linhas)")
        del df_alerta_xlsx  # ✅ LIBERA MEMÓRIA

    del df_regras_auditoria  # ✅ LIBERA MEMÓRIA

    # ────────────────────────────────────────────────────────────────────────
    # 7. CONSOLIDAÇÃO (lendo dos parquets temporários)
    # ────────────────────────────────────────────────────────────────────────
    print("\n  ─── CONSOLIDAÇÃO (lendo parquets temporários) ───")

    # IMPORTANTE: o todas_regras_auditoria.parquet é a soma das regras já
    # separadas em regra1..regra5 — incluí-lo na concatenação duplicaria tudo.
    # Por isso ele é EXCLUÍDO da consolidação.
    parquet_files = sorted(
        f for f in TEMP_DIR.glob("*.parquet")
        if f.name != "todas_regras_auditoria.parquet"
    )
    print(f"  → Consolidando {len(parquet_files)} arquivos...")

    resultados = pd.concat(
        [pd.read_parquet(f) for f in parquet_files],
        ignore_index=True
    )

    print(f"  → Total consolidado: {len(resultados):,} alertas")
    if not resultados.empty and "id_aluno" in resultados.columns:
        print(f"     ({resultados['id_aluno'].nunique():,} alunos únicos)")

    # 8. Monta resumo
    print("\n  ─── RESUMO ───")
    contagem = (
        resultados.groupby("alerta").size().reset_index(name="Qtd")
        if not resultados.empty
        else pd.DataFrame(columns=["alerta", "Qtd"])
    )
    resumo = (
        pd.DataFrame({"alerta": TODOS_ALERTAS})
        .merge(contagem, on="alerta", how="left")
        .fillna({"Qtd": 0})
        .astype({"Qtd": int})
    )

    if not contagem.empty:
        print("\n  📊 Breakdown por tipo de alerta:")
        for _, row in contagem.sort_values("Qtd", ascending=False).iterrows():
            print(f"     • {row['alerta']:<50} {row['Qtd']:>8,}")

    # 9. Salva parquets finais
    print("\n  ─── SALVANDO ARQUIVOS ───")
    print("  → Salvando parquets finais...")
    resultados.to_parquet(PARQUET_RESULTADOS, index=False)
    resumo.to_parquet(PARQUET_RESUMO, index=False)
    print(f"     ✅ resultados.parquet → {len(resultados):,} alertas")
    print(f"     ✅ resumo.parquet     → {len(resumo)} tipos de alerta")

    # 10. Atualiza Excel (passo cosmético — não deve derrubar o pipeline).
    #     Os parquets acima são o que o app consome; o xlsx só alimenta gráficos.
    print("  → Atualizando dados_graficos.xlsx...")
    try:
        _salvar_excel(resumo)
        print("     ✅ dados_graficos.xlsx atualizado")
    except Exception as e:
        print(f"     ⚠️  Falha ao atualizar o Excel (parquets já foram salvos): {e}")
        print("        Os resultados estão salvos; o pipeline continua normalmente.")

    # ✅ LIMPA ARQUIVOS TEMPORÁRIOS (só agora, com tudo consolidado com sucesso)
    # Em pastas do OneDrive, o cliente de sincronização pode manter um lock na
    # pasta/arquivos. A remoção é cosmética — não deve derrubar o pipeline.
    print("  → Limpando arquivos temporários...")
    import time

    # Remove os parquets um a um (ignora os que estiverem travados)
    for f in TEMP_DIR.glob("*.parquet"):
        for tentativa in range(3):
            try:
                f.unlink()
                break
            except PermissionError:
                time.sleep(0.5)  # dá tempo do OneDrive soltar o lock
        else:
            print(f"     ⚠️  Não consegui remover {f.name} (travado pelo OneDrive?).")

    # Remove a pasta (só funciona se estiver vazia; tenta algumas vezes)
    pasta_removida = False
    for tentativa in range(3):
        try:
            TEMP_DIR.rmdir()
            pasta_removida = True
            break
        except OSError:
            time.sleep(0.5)

    if pasta_removida:
        print("     ✅ Temporários removidos")
    else:
        print("     ⚠️  Parquets removidos, mas a pasta 'temp' ficou (lock do OneDrive).")
        print("        Sem problema: a próxima execução reaproveita a pasta.")

    fim = datetime.now()
    duracao = (fim - inicio).total_seconds()
    print(f"\n[{fim:%H:%M:%S}] ✅ Processamento concluído em {duracao:.1f}s.")


if __name__ == "__main__":
    try:
        main()
    except Exception as e:
        print(f"\n❌ ERRO FATAL: {e}")
        import traceback
        traceback.print_exc()
        raise
